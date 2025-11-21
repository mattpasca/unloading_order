"""
Route planner for the transports of Gorini Piante.
The user gives input through the 'Costi.xlsx' file: names of the companies we have to deliver to 
(which usually has a good matching with 'Acronimo' in the customer database). The program retrieves
detailed address information and decides optimal unloading order.
A mail for marta transport is printed automatically. The route is printed in an html file.

TODO
    - handle all errors (nan coordinates, invalid keys etc)
    - handle unloading order in mail for unmatched customers
    - clean up code
    - log all funtion calls with trz except blocks
"""

import os
from pathlib import Path
import logging
import sys
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
import Levenshtein as Lev
import pgeocode as pg
import requests
from geopy.geocoders import Nominatim
import time
import folium
import polyline
import numpy as np
import json

output_dir = Path("output")
output_dir.mkdir(exist_ok=True)  # Create the directory if it doesn't exist
# Define the output file path
output_filename = output_dir / "Costi_compilato.xlsx"

log_dir = Path("logs")
log_dir.mkdir(exist_ok=True)

log_filename = log_dir / f"script_log_{datetime.now().strftime('%Y-%m-%d')}.log"

logging.basicConfig(
    filename=log_filename,
    filemode="a",
    format="%(asctime)s [%(levelname)s] %(message)s",
    level=logging.INFO,
)
logging.getLogger().addHandler(logging.StreamHandler(sys.stdout))
 
base_path = Path(__file__).parent

common_terms = {
    1: 'baumschule',
    2: 'galabau',
    3: 'garten',
    4: 'gartencenter',
    5: 'baumschulen',
    6: 'gartenbau',
    7: 'landschaftsbau',
    8: 'gmbh',
    9: 'ag',
    10: 'rosen',
    11: 'boerse',
    12: 'blumenboerse',
    13: 'blumen',
    14: 'gbr',
    15: 'kg',
    16: 'gartenarchitektur',
    17: 'gartendesign'
}

customer_db = Path("Z - Dati") / "CLIENTI_FAT.csv"
customer_info = pd.read_csv(customer_db, sep=';', dtype=str, keep_default_na=False)
customer_info = customer_info.fillna('')

wb = load_workbook("Costi.xlsx")
ws = wb["Foglio1"]

def autocomplete(customers, xl_file):
    # Map column labels -> Excel 1-based index
    col_index = {col: j+1 for j, col in enumerate(xl_file.columns)}
    for i, row in xl_file.iterrows():
        cust_data = customers.get(row['Cliente'], {})
        for col, value in cust_data.items():
            if col in col_index:  # only fill if column exists in sheet
                ws.cell(row=i+2, column=col_index[col], value=value)
    wb.save(output_filename)

# read data of currently palnned transport
def	read_transport_plan(xl_file):
    names = []
    names = xl_file['Cliente'].tolist()
    for name in names:
        if name == '':
            names.remove(name)
    return (names)

# remove common terms from customer's name and measure Levenshtein ratio
def	fuzzy_match(str1, str2):
    str1 = str1.lower().strip()
    str2 = str2.lower().strip()
    for key in common_terms:
        str1 = str1.replace(common_terms[key], '')
        str2 = str2.replace(common_terms[key], '')
    return (Lev.ratio(str1.strip(), str2.strip()))

# retrieve adress in CSV from name
def	get_adress(name, customer_info, xl_file):
    search_fields = {
        1: 'Acronimo',
        2: 'Ragione Sociale 1',
        3: 'Ragione Sociale 2'
    }
    customer_adress = {}
    try: # if Indirizzo column is not there...
        manual_add = str(xl_file[xl_file['Cliente']==name]['Indirizzo'].iloc[0])
        if manual_add != '' and '-' in manual_add:
            naz, cap = manual_add.split('-', 1)
            customer_adress['Country'] = naz
            customer_adress['CAP'] = cap.strip()
            print(f"Indirizzo di {name} inserito manualmente\n")
            return (customer_adress)
    except:
        pass
    try:
        name = name.lower()
    except:
        print(Exception)
    max_similarity = 0
    for i, row in customer_info.iterrows():
        for key in search_fields:
            current_ratio = fuzzy_match(name, row[search_fields[key]].strip())
            if current_ratio > 0.79 and current_ratio > max_similarity:
                max_similarity = current_ratio
                customer_adress['Codice'] = f"{row['Naz']}" + '/' + f"{row['Codice']}"
                customer_adress['CAP'] = row['CAP'].strip()
                customer_adress['City']= row['Localita\''].strip()
                customer_adress['Country'] = row['Nazione'].strip()
                customer_adress['Ragione Sociale'] = row['Ragione Sociale 1'].strip() + row['Ragione Sociale 2'].strip()
                customer_adress['Indirizzo'] = customer_adress['Country'] + '-' + customer_adress['CAP'] + ' ' + row['Indirizzo'].strip()
                break
    if max_similarity == 0:
        print(f"Indirizzo di {name} non trovato! Inserire manualmente\n")
    else:
        print(f'Trovato indirizzo di: {name}\n')
    return (customer_adress)

# we need coordinates for the trip call to OSM server
def get_coordinates(zip_code, country):
    coordinates = {}
    geodb = pg.Nominatim(country)
    coordinates['latitude'] = geodb.query_postal_code(zip_code)['latitude']
    coordinates['longitude'] = geodb.query_postal_code(zip_code)['longitude']
    return (coordinates)

# unnecessary
def get_zip_code(coord):
    geolocator = Nominatim(user_agent="matteo.pascale@gorinipiante.it", timeout=10)
    location = geolocator.reverse((coord['latitude'], coord[0]), exactly_one=True)
    if location and 'postcode' in location.raw['address']:
        return location.raw['address']['postcode']
    return None

# build a dictionary of the customers with: name, address, zip, coordinates
def	customer_dict(xl_file):
    customers = {}
    print('Analizzo file excel\n')
    for name in read_transport_plan(xl_file):
        if name != '':
            customers[name] = get_adress(name, customer_info, xl_file)
            if customers[name] and customers[name]['Country']:
                customers[name]['Coordinates'] = get_coordinates(customers[name]['CAP'], customers[name]['Country'].strip())
            time.sleep(2)
    return (customers)

def google_maps(unloading_order, adress_dict):
    base_url = "https://www.google.com/maps/dir/"
    waypoint_list = []
    for i in range (1, len(unloading_order) + 1):
        name = unloading_order[i]
        lat = adress_dict[name]['Coordinates']['longitude']
        lng = adress_dict[name]['Coordinates']['latitude']
        waypoint_list.append(f"{lng},{lat}")
    waypoints = "/".join(waypoint_list)
    google_maps_url = base_url + waypoints
    maps_link = output_dir / "google.txt"
    with open(maps_link, "w") as f:
        f.write(google_maps_url)


# api call. Note source=first & roundtrip=false & geometries=polyline
def	osm_request(coord_list):
    coord_str = ";".join([f"{lon},{lat}" for lon, lat in coord_list])
    url = (
        f"http://router.project-osrm.org/trip/v1/driving/{coord_str}"
        "?source=first&roundtrip=false&geometries=polyline6&overview=full"
    )
    print(url)
    r = requests.get(url)
    data = r.json()
    return (data)

# get the polyline and the unloading order. The latter is constructed with waypoint indices
def shortest_path(adress_dict):
    coord_list = []
    coord_list.append((10.972482955970866, 43.91835625149449)) # Chiazzano 51100
    for customer in adress_dict:
        try:
            coord = (adress_dict[customer]['Coordinates']['longitude'], adress_dict[customer]['Coordinates']['latitude'])
            if all(val == val for val in coord):
                coord_list.append(coord)
            else:
                print(f"Coordinate di {adress_dict[customer]} non trovate!")
        except:
            continue
    osm_response = osm_request(coord_list)
    unloading_order = {}
    for wp, customer in zip(osm_response["waypoints"][1:], adress_dict.keys()):
        unloading_order[wp["waypoint_index"]] = customer
    print(adress_dict)
    print(unloading_order)
    google_maps(unloading_order, adress_dict)
    return (unloading_order, osm_response)

# standard formulation for the request to Marta Transport    
def print_email(customer_dict, xl_file, unloading_order):
    print(f"CUSTOMER DICT {customer_dict}")
    text = """Camion <DATA>\n
    Cabina:\n"""
    l = len(unloading_order)
    for i in range(0, len(unloading_order)):
        name = unloading_order[l-i]
        if name != '':
            try:
                text += f"{customer_dict[name]['Country']}-{customer_dict[name]['CAP']}      {name}      ordine:\n"
            except:
                text += f"X-1234    {name}      ordine:\n"
    mail_file = output_dir / "lista_di_carico.txt"
    with open(mail_file, "w") as f:
        f.write(text)
    print('Testo mail stampato in lista_di_carico.txt\n')

def transit_time(distances):
    transit_times = []
    break_count = 0
    shift_count = 0    
    for stop in distances:
        t_time = stop/60 # average 60 Km/h
        break_count = round(t_time/4.5, 0)
        shift_count = round(t_time/9, 0)
        t_time += break_count * 0.75 + shift_count * 11
        t_time = round(t_time, 2)
        transit_times.append(t_time)
    return (transit_times)


def print_route(osm_result, unloading_order):
    # Decode geometry
    coords = polyline.decode(osm_result["trips"][0]["geometry"], precision=6)
    m = folium.Map(location=coords[0], zoom_start=8)

    # Draw route polyline
    folium.PolyLine(locations=coords, color="blue", weight=5, opacity=0.7).add_to(m)

    # Extract distances
    total_distance = osm_result["trips"][0]["distance"] / 1000  # meters → km
    legs = osm_result["trips"][0]["legs"]

     # --- Add stop markers ---
    waypoints = osm_result["waypoints"]
    for i, wp in enumerate(waypoints):
        if i == 0:
            stop_name = "Gorini"
        else:
            key = wp["waypoint_index"] 
            stop_name = unloading_order.get(key, f"Stop {i}")
        folium.Marker(
            location=[wp["location"][1], wp["location"][0]],  # lon,lat → lat,lon
            popup=f"{i+1}. {stop_name}",
            tooltip=stop_name,
            icon=folium.Icon(color="red" if i == 0 else "green", icon="info-sign")
        ).add_to(m)

    # Build HTML table
    table_html = f"""
    <div style="background:white; padding:10px; border:2px solid black; max-height:200px; overflow:auto;">
    <h4>Percorso</h4>
    <b>Totale:</b> {total_distance:.1f} km
    <table border="1" style="border-collapse:collapse; margin-top:5px; width: 100%;">
        <tr>
            <th style="width:150px">Fermata</th>
            <th style="width:100px">Distanza (km)</th>
            <th style="width:100px">Tempo (h)</th>
        </tr>
    """

    print(f"Number of legs: {len(legs)}")
    print(f"Number of slices: {len(unloading_order)}")

    transit_times = transit_time([leg["distance"] / 1000 for leg in legs])
    json_dict = []
    for i, leg in enumerate(legs):
        for j in unloading_order.keys():
            if j == i + 1:
                name = unloading_order[j]
        dist_km = leg["distance"] / 1000
        table_html += f"<tr><td>{i+1}. {name}</td><td>{dist_km:.1f}</td><td>{transit_times[i]}</td></tr>"
        json_dict.append({"name": name, "hours_from_prev": transit_times[i]})
    table_html += "</table></div>"
    run_data = {
        "id": "",
        "departure": "",
        "image": "001.png",
        "customers": json_dict,
    }
    output_path = output_dir / "tabellone.json"
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(run_data, f, indent=2)

    # Add as FloatImage overlay (fixed on map)
    from branca.element import Figure, MacroElement, Template
    template = Template("""
        {% macro html(this, kwargs) %}
        """ + table_html + """
        {% endmacro %}
    """)
    macro = MacroElement()
    macro._template = template
    m.get_root().add_child(macro)

    # Save
    route_filename = output_dir / "percorso.html"
    m.save(route_filename)
    print('Percorso salvato nel file percorso.html\n')

from PIL import Image, ImageDraw, ImageFont

# ChatGPT fct
def fit_text_in_box(customers, step, draw, font_path="arial.ttf", max_font_size=200):
    ll = (233,250)
    ul = (233, 96)
    min_font_size = max_font_size
    for name in customers:
        text = customers[name]
        font_size = max_font_size
        font = ImageFont.truetype(font_path, font_size)

        # Decrease font size until text fits
        while font_size > 15:
            # get text bounding box (x0,y0,x1,y1)
            bbox = draw.multiline_textbbox((0,0), text, font=font)
            text_width, text_height = bbox[2] - bbox[0], bbox[3] - bbox[1]
            box_height = abs(ll[1] - ul[1])
            if text_width <= step and text_height <= box_height/3:
                break
            font_size -= 1
        if font_size <= min_font_size:
            min_font_size = font_size
    font = ImageFont.truetype(font_path, font_size)
    return font  # returns smallest font if nothing fits

def draw_lines(step, num, draw):
    ll = (233,250)
    ul = (233, 96)
    for i in range(1, num):
        point1 = (ul[0]+i*step, ul[1])
        point2 = (ll[0]+i*step, 400)
        draw.line([point1, point2], fill="black", width=3)
    return (draw)

def draw_truck(unloading_order, customers):
    # Semi trailer pixel coordinates
    ll = (233,250)
    ul = (233, 96)
    ur = (938, 96)
    lr = (938, 250)
    width = lr[0] - ll[0] 
    height = ul[1] - ll[1]
    name_pos = (ul[0], ul[1] - height/4)
    img_path = Path("Z - Dati") / "truck.png"
    img = Image.open(img_path)
    draw = ImageDraw.Draw(img)
    font = ImageFont.truetype("arial.ttf", 20)
    step = width/len(unloading_order)
    draw = draw_lines(step, len(unloading_order), draw)
    x, y = name_pos
    font = fit_text_in_box(unloading_order, step, draw)
    for i in reversed(range(len(unloading_order))):
        txt = unloading_order[i+1].replace("\\n", "\n")
        try:
            txt += "\n" + customers[txt]['Codice']
        except:
            pass
        draw.multiline_text((x,y), txt, font=font, fill="black", align="center", spacing=2)
        x += step
        
    img_filename = output_dir / "truck_loaded.png"
    img.save(img_filename)    

def	main():
    xl_file = pd.read_excel('Costi.xlsx')
    xl_file = xl_file.fillna('') # empty cells are treated as NaN (float). this leads to errors with string methods
    customers = customer_dict(xl_file)
    unloading_order, osm_result = shortest_path(customers)
    autocomplete(customers, xl_file)
    print_email(customers, xl_file, unloading_order)
    print_route(osm_result, unloading_order)
    draw_truck(unloading_order, customers)

if __name__ == "__main__":
    try:
        main()
    except Exception:
        logging.exception("Chiama Matteo")
        sys.exit(1)
